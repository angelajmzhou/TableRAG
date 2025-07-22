import sys
from openpyxl import load_workbook
import torch
from typing import Union, List, Tuple
from transformers import AutoTokenizer, AutoModel, AutoModelForSequenceClassification
from tqdm import tqdm
import numpy as np

def sigmoid(x) :
    return 1 / (1 + np.exp(-x))

class Embedder :
    def __init__(self, model_path, device_id=0) -> None:
        self.model_path = model_path
        self.device = torch.device(f"cuda:{device_id}")
        self.tokenizer = AutoTokenizer.from_pretrained(self.model_path, use_fast=True)
        self.model = AutoModel.from_pretrained(self.model_path)
        self.model = self.model.to(self.device)
        self.model.eval()

    @torch.no_grad()
    def encode(self, texts, batch_size=64, max_length=512):
        all_embeddings = []
        i = 0
        current_batch_size = batch_size

        while i < len(texts):
            error_count = 0

            while True:
                try:
                    batch = texts[i : i+current_batch_size]

                    features = self.tokenizer(
                        batch,
                        padding=True,
                        truncation=True,
                        max_length=max_length,
                        return_tensors="pt"
                    ).to(self.device)
                    #[last_hidden_state, hidden_states (optional), attentions (optional)]

                    model_output = self.model(**features)
                    
                    #of last_hidden_state: (batch_size, sequence_length, hidden_dim))
                    cls_embs = model_output[0][:, 0].cpu().numpy()

                    all_embeddings.extend(cls_embs)
                    torch.cuda.empty_cache()
                    i += current_batch_size
                    break  # success

                except (torch.cuda.OutOfMemoryError, RuntimeError) as e:
                    torch.cuda.empty_cache()
                    current_batch_size = max(1, current_batch_size // 2)
                    error_count += 1
                    print(f"[OOM] Retrying batch with size {current_batch_size}")
                    if error_count > 5:
                        raise RuntimeError("Too many OOM retries in encode()")

        return all_embeddings
class Reranker :
    def __init__(
        self,
        model_name_or_path: str = None,
        use_fp16: bool = False,
        inference_mode: str = "huggingface",
        cache_dir: str = None,
        device: Union[str, int] = 4
    ) -> None:
        
        self.inference_mode = inference_mode
        self.tokenizer = AutoTokenizer.from_pretrained(model_name_or_path, cache_dir=cache_dir)

        if device and isinstance(device, str) :
            self.device = torch.device(device)
            if device == "auto" :
                use_fp16 = False
        else :
            if torch.cuda.is_available() :
                if device is not None :
                    self.device = torch.device(f"cuda:{0}")
                else :
                    self.device = torch.device("cuda")       

        self.model = AutoModelForSequenceClassification.from_pretrained(
            model_name_or_path,
            cache_dir=cache_dir,
            trust_remote_code=True
        )
        if use_fp16 :
            self.model.half()
        self.model.eval()

        self.model = self.model.to(self.device)

        if device is None :
            self.num_gpus = torch.cuda.device_count()
            if self.num_gpus > 1 :
                print(f"----- using {self.num_gpus}*GPUs -------")
                self.model = torch.nn.DataParallel(self.model)
        else :
            self.num_gpus = 1

    @torch.no_grad()
    def compute_score(self, sentence_paris: Union[List[Tuple[str, str]], Tuple[str, str]], batch_size: int = 64,
                        max_length: int = 512, normalize: bool = False) -> List[float] :
        if self.num_gpus > 0 :
            batch_size = batch_size * self.num_gpus

        assert isinstance(sentence_paris, list)
        if isinstance(sentence_paris[0], str) :
            sentence_paris = [sentence_paris]

        all_scores = []
        i = 0

        while i < len(sentence_paris):
            current_batch_size = batch_size
            error_count = 0
            while True:
                try:
                    batch = sentence_paris[i:i + current_batch_size]
                    inputs = self.tokenizer(
                        batch,
                        padding=True,
                        truncation=True,
                        max_length=max_length,
                        return_tensors="pt"
                    ).to(self.device)

                    scores = self.model(**inputs, return_dict=True).logits.view(-1).float()
                    all_scores.extend(scores.cpu().numpy().tolist())
                    torch.cuda.empty_cache()
                    i += current_batch_size
                    break  # exit retry loop if successful

                except (torch.cuda.OutOfMemoryError, RuntimeError) as e:
                    current_batch_size = max(1, current_batch_size // 2)
                    error_count += 1
                    print(f"[Retry] Reducing batch size to {current_batch_size}")
                    torch.cuda.empty_cache()

                    if error_count > 5:
                        raise RuntimeError("Exceeded maximum OOM retries")

        if normalize:
            all_scores = [sigmoid(score) for score in all_scores]

        return all_scores

#used to convert excel files to markdown tables
def excel_to_markdown(file_path) :
    workbook = load_workbook(file_path)

    content = ""
    file_name = file_path.split("/")[-1]
    table_name = file_name.replace(".xlsx", "")
    content += f"Table name: {table_name}\n"
    for sheet_name in workbook.sheetnames :
        work_sheet = workbook[sheet_name]
        row_count = 0
        for i, row in enumerate(work_sheet) :
            columns = []
            for column in row :
                if column.value is None :
                    continue
                columns.append(column.value)

            content += " | " + " | ".join(columns) + " | \n"
            if i == 0 :
                content += " | " + " | ".join(["---"]*len(columns)) + " | \n"
            row_count += 1
    return content

#for testing functionality
if __name__ == '__main__' :
    print(excel_to_markdown("./test.xlsx"))